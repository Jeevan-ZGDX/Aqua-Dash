"""Serialization of report datasets to CSV, Excel and PDF."""

import csv
import io
from typing import Iterable

from app.core.exceptions import ReportGenerationError


def to_csv(headers: list[str], rows: list[list]) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    return buffer.getvalue().encode("utf-8-sig")


def to_excel(headers: list[str], rows: list[list], sheet_name: str = "Report") -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise ReportGenerationError("Excel export requires openpyxl.") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name[:31]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")

    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    for row in rows:
        sheet.append(list(row))

    for idx, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(idx)].width = max(12, min(len(str(header)) + 4, 40))

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def to_pdf(title: str, headers: list[str], rows: list[list]) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import landscape, A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    except ImportError as exc:  # pragma: no cover
        raise ReportGenerationError("PDF export requires reportlab.") from exc

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), leftMargin=10 * mm, rightMargin=10 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    story = [Paragraph(title, styles["Title"]), Spacer(1, 6 * mm)]

    data: list[list] = [headers]
    data.extend([_cell(v) for v in row] for row in rows)
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E78")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EAF1F8")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ]
        )
    )
    story.append(table)
    doc.build(story)
    return buffer.getvalue()


def _cell(value) -> str:  # noqa: ANN001
    if value is None:
        return ""
    return str(value)
